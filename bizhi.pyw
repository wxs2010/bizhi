import ctypes
import json
import os
import re
import time
from datetime import datetime

import requests
from PIL import Image, ImageDraw, ImageFont
from bs4 import BeautifulSoup
from openpyxl import load_workbook


# 定义替换函数
# 获取列字母
def get_column_letter(column_number):
    return chr(ord('A') + column_number)

def fetch_and_parse_weather(url):
    """
    从指定URL获取天气信息并解析。

    该函数发送一个HTTP GET请求以获取天气数据网页，然后使用正则表达式从JavaScript代码中提取小时预报数据，
    解析并转换为更易处理的字典格式。

    参数:
    url (str): 包含天气信息的网页URL。

    返回:
    list: 包含小时预报信息的字典列表。如果请求失败或未找到数据，返回 "error"。
    """
    # 发送HTTP GET请求
    try:
        response = requests.get(url)
        response.encoding = 'utf-8'

        # 检查请求是否成功
        if response.status_code == 200:
            # 获取网页内容
            html_content = response.text
            #print(html_content)

            # 使用正则表达式匹配JavaScript中的forecast_1h数组
            match = re.search(r'var forecast_1h = (\[.*?\]);', html_content, re.DOTALL)

            if match:
                # 将匹配到的JSON字符串转换成Python列表
                hourly_forecast = json.loads(match.group(1))

                # 初始化结果列表
                results = []
                for forecast in hourly_forecast:
                    # 解析天气预报中的各项数据
                    time = forecast['time']
                    weather = forecast['weather']
                    temp = forecast['temp']
                    wind_level = forecast['windL']
                    wind_direction = forecast['windD']

                    # 构造结果字典
                    result = {
                        "Time": f"{time}",
                        "Weather": weather,
                        "Temperature": f"{temp}°C",
                        "Wind": f"{wind_direction}\n{wind_level} "
                    }
                    # 将结果字典添加到结果列表中
                    results.append(result)

                # 返回结果列表
                return results
            else:
                # 如果没有找到天气信息，返回 "error"
                return "error"
        else:
            # 如果请求失败，返回 "error"
            return "error"
    except:
        return "error"

def fetch_and_parse_weather2(url):
    try:
        # 发送HTTP GET请求
        response = requests.get(url)
        response.encoding = 'utf-8'

        # 检查请求是否成功
        if response.status_code == 200:
            # 获取网页内容
            html_content = response.text

            # 使用BeautifulSoup解析HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            items = soup.find_all('li', class_='blue-item')

            # 遍历每个列表项并提取天气信息
            weather_results = []
            for item in items:
                # 提取天气状况
                weather_info = item.find('p', class_='weather-info').get_text().strip()

                # 使用正则表达式移除非天气描述部分
                weather_info = re.sub(r'[^\u4e00-\u9fa5]+级', '', weather_info)

                # 将信息添加到字典中
                weather_results.append({
                    "weather_info": weather_info
                })

            # 使用正则表达式匹配JavaScript中的eventDay和eventNight数组
            day_match = re.search(r'var eventDay = (\[.*?\]);', html_content, re.DOTALL)
            night_match = re.search(r'var eventNight = (\[.*?\]);', html_content, re.DOTALL)

            if day_match and night_match:
                # 将匹配到的JSON字符串转换成Python列表
                event_day = json.loads(day_match.group(1))
                event_night = json.loads(night_match.group(1))

                # 提取日期信息
                date_container = soup.find('ul', class_='date-container')
                date_items = date_container.find_all('li', class_='date-item') if date_container else []

                # 存储结果的列表
                results = []
                for i, date_item in enumerate(date_items):
                    date = date_item.find('p', class_='date').get_text()
                    date_info = date_item.find('p', class_='date-info').get_text()

                    # 替换“星期一”、“星期二”等为“周一”、“周二”等
                    date_info = re.sub(r'星期', r'周', date_info)

                    day_temp = event_day[i]
                    night_temp = event_night[i]

                    # 将信息添加到字典中
                    results.append({
                        "date": date,
                        "date_info": date_info,
                        "day_temp": day_temp,
                        "night_temp": night_temp
                    })

                # 合并天气信息和日期信息
                combined_results = []
                for weather, result in zip(weather_results, results):
                    combined_results.append({
                        "date": result["date"],
                        "date_info": result["date_info"],
                        "day_temp": result["day_temp"],
                        "night_temp": result["night_temp"],
                        "weather_info": weather["weather_info"]
                    })

                return combined_results  # 返回字典列表
            else:
                return "error"  # 未找到天气信息
        else:
            return "error"  # 请求失败
    except:
        return "error"

def read_all_items_from_column(file_path, sheet_name, column_number):
    # 加载工作簿
    workbook = load_workbook(filename=file_path, data_only=True)

    # 获取指定的工作表
    sheet = workbook[sheet_name]

    # 读取指定列的数据
    results = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column_number, max_col=column_number):
        for cell in row:
            results.append(cell.value)

    return results

# 计算距离高考的时间
def gaokao(year):
    now = datetime.now()
    gaokao_date = datetime(year, 6, 7, 9, 0)
    time_difference = gaokao_date - now
    if time_difference.days < 0:
        gaokao_date = datetime(year + 1, 6, 7, 9, 0)
        time_difference = gaokao_date - now
    days, seconds = time_difference.days, time_difference.seconds
    hours = (seconds + 60) // 3600
    minutes = (seconds % 3600 + 60) // 60 % 60
    return f"距离高考还有\n{days}天{hours}时{minutes}分"

# 在图片上添加文字
def add_text_to_image(image_path, text, data_list, output_path):
    # 打开图片文件
    with Image.open(image_path) as img:
        # 创建一个可以在给定图像上绘图的对象
        draw = ImageDraw.Draw(img)

        # 字体的格式
        font = ImageFont.truetype("simkai.ttf", 100)

        # 在指定位置写入文字
        text_position = (2600, 50)
        draw.text(text_position, text, font=font, fill=(255, 255, 255))

        j = -70
        for i, item in enumerate(data_list):
            if item is None:
                j += 30
            else:
                j += 100
                draw.text((3550 - len(item) * 50, j), item, font=font, fill=(255, 255, 255))
        if hwea != "error":
            j = 1400
            font = ImageFont.truetype("simkai.ttf", 50)
            now = int(datetime.now().hour) # 17
            tn = int(float(hwea[0]['Time'])) # 8
            if now > tn:
                for hw in range(now - tn, now - tn + 6):
                    j += 150
                    draw.text((j, 80), str(hwea[hw]['Time'] + ":00\n" + hwea[hw]['Weather'] + "\n" + hwea[hw]['Temperature']), font=font, fill=(255, 255, 255))
            else:
                for hw in range(6):
                    j += 150
                    draw.text((j, 80), str(hwea[hw]['Time'] + ":00\n" + hwea[hw]['Weather'] + "\n" + hwea[hw]['Temperature']), font=font, fill=(255, 255, 255))

        if wea != "error":
            j = 1290
            font = ImageFont.truetype("simkai.ttf", 50)
            for w in range(7):
                j += 260
                draw.text((j, 230), str(wea[w]['date'] + " " + wea[w]['date_info'] + "\n" + wea[w]['weather_info'] + "\n" + str(wea[w]['day_temp']) + "℃/" + str(wea[w]['night_temp']) + "℃"), font=font, fill=(255, 255, 255))
        font = ImageFont.truetype("simkai.ttf", 80)

        # 在指定位置写入文字
        font = ImageFont.truetype("simkai.ttf", 75)
        draw.text((2800, 500), upcoming_birthdays, font=font, fill=(255, 255, 255))

    img.save('data/temp.jpg', format='JPEG')


def find_upcoming_birthdays(file_name):
    # 获取今天的日期
    today = datetime.now().date()

    # 加载Excel文件
    wb = load_workbook(filename=file_name)
    ws = wb.active  # 选择活动的工作表，默认为第一个工作表

    # 存储每个人的生日信息
    birthdays = []

    # 遍历工作表中的每一行
    for row in ws.iter_rows(values_only=True):
        if row:
            name = row[0]  # 假设姓名在第一列
            birth_date_str = row[1]  # 假设生日在第二列且为字符串形式 "MMDD"
            if name and birth_date_str:  # 检查是否有值
                month = int(birth_date_str[:2])
                day = int(birth_date_str[2:])

                # 创建一个生日日期对象
                try:
                    birthday_this_year = datetime(today.year, month, day).date()

                    # 如果生日在今年已经过去，则计算明年的生日日期
                    if birthday_this_year < today:
                        birthday_this_year = datetime(today.year + 1, month, day).date()

                    birthdays.append((name, birthday_this_year))
                except ValueError:
                    print(f"Illegal date for {name}: {birth_date_str}")


    # 根据生日日期排序
    birthdays.sort(key=lambda x: x[1])

    # 创建一个描述生日的字符串列表
    birthday_descriptions = []
    for name, bd in birthdays:
        # 如果找到了三个生日，就停止查找
        if len(birthday_descriptions) >= 3:
            break

        days_until_birthday = (bd - today).days
        if days_until_birthday > 0:
            description = f"{name}：{days_until_birthday}天"
        elif days_until_birthday < 0:
            # 如果生日已过，跳过这个人
            continue
        else:
            description = f"{name}：今天"

        birthday_descriptions.append(description)

    # 添加标题
    birthday_descriptions.insert(0, "近期生日：")

    # 返回格式化的字符串
    return '\n'.join(birthday_descriptions)
# 更改壁纸
def change_wallpaper(image_path):
    print("change")
    SPI_SETDESKWALLPAPER = 20
    ctypes.windll.user32.SystemParametersInfoW(SPI_SETDESKWALLPAPER, 0, image_path, 3)
# 主函数
def main():
    global hwea, wea, upcoming_birthdays
    output_path = 'data/temp.jpg'
    image_path = 'data/seewo.jpg'

    # 更新壁纸
    def change():
        add_text_to_image(image_path, gaokao(2028), data_list, output_path)
        wallpaper_path = os.path.join(os.getcwd(), output_path)
        if os.path.exists(wallpaper_path):
            change_wallpaper(wallpaper_path)
        else:
            print("Wallpaper file not found.")

    upcoming_birthdays = find_upcoming_birthdays('data/BD.xlsx')
    # 初始更新
    hwea = fetch_and_parse_weather(url)
    wea = fetch_and_parse_weather2(url2)
    # 读取当天课程表
    data_list = read_all_items_from_column('data/课程表.xlsx', 'Sheet1', datetime.now().weekday() + 1)
    change()

    while True:
        upcoming_birthdays = find_upcoming_birthdays('data/BD.xlsx')
        hwea = fetch_and_parse_weather(url)
        wea = fetch_and_parse_weather2(url2)
        # 读取当天课程表
        data_list = read_all_items_from_column('data/课程表.xlsx', 'Sheet1', datetime.now().weekday() + 1)
        for i in range(10):
            change()
            time.sleep(60 - datetime.now().second)

if __name__ == "__main__":
    while True:
        try:
            print(find_upcoming_birthdays('data/BD.xlsx'))
            url2 = "https://forecast.weather.com.cn/town/weathern/101280206.shtml"
            url = "https://forecast.weather.com.cn/town/weather1dn/101280206.shtml"

            #  https://www.weather.com.cn/weather1dn/101280206.shtml
            hwea = fetch_and_parse_weather(url)
            wea = fetch_and_parse_weather2(url2)
            print(hwea)
            print(wea)
            main()
        except:
            time.sleep(60)
